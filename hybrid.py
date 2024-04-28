from PyQt5 import uic
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QButtonGroup, QErrorMessage, QMessageBox
from PyQt5.QtGui import QPixmap
import win32com.client as com
import os
from openpyxl import load_workbook
import shutil
import warnings
from writing import writing_campo
from openpyxl import load_workbook
from hybrid_ui import Ui_MainWindow
from model import *
import xlwings as xw

warnings.filterwarnings("ignore", category=DeprecationWarning)
nombres_vehiculos = []
enviado_contador = 1
param_groups = {}

class MiVentana(QMainWindow):
    def __init__(self): #Ready
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        ####IMAGES####
        imagen = QPixmap('./images/car_follow.png')
        self.ui.label.setPixmap(imagen)
        imagen_2 = QPixmap('./images/lane_change.png')
        self.ui.label_2.setPixmap(imagen_2)
        image_3 = QPixmap('./images/lateral_behave.png')
        self.ui.label_3.setPixmap(image_3)
        image_4 = QPixmap('./images/logo.png')
        self.ui.label_5.setPixmap(image_4)

        #BOTONES

        self.ui.start.clicked.connect(self.ejecutar_programa)
        self.ui.carpet.clicked.connect(self.carpeta)
        self.ui.report.clicked.connect(self.data_model)
        self.ui.activar.clicked.connect(self.data_campo)
        self.ui.liviano.clicked.connect(self.livianos)
        self.ui.menor.clicked.connect(self.menores)
        self.ui.publico.clicked.connect(self.publicos)
        self.ui.carga.clicked.connect(self.cargas)
        self.ui.fijar.clicked.connect(self.fijars)
        self.ui.exportar.clicked.connect(self.export_params_2_excel)
        self.ui.get_pushButton.clicked.connect(self.get)
        self.ui.run.clicked.connect(self.run_vissim)
        
        #BOTONES PARA LOS TURNOS
        button_group = QButtonGroup(self)
        button_group.addButton(self.ui.early)
        button_group.addButton(self.ui.morning)
        button_group.addButton(self.ui.evening)
        button_group.addButton(self.ui.night)
        button_group.setExclusive(True)

        #BOTONES PARA LAS VERSIONES
        button_group_2 = QButtonGroup(self)
        button_group_2.addButton(self.ui.checkBox)
        button_group_2.addButton(self.ui.checkBox_2)
        button_group_2.addButton(self.ui.checkBox_3)

    def carpeta(self): #Ready
        self.path_file, self.inpx_name = QFileDialog.getOpenFileName(self, "Seleccionar Archivo .inpx","c:\\","Archivos .inpx (*.inpx)")

    def data_campo(self): #Ready
        #Checking buttons:
        while not (
            self.ui.early.isChecked() or
            self.ui.morning.isChecked() or
            self.ui.evening.isChecked() or
            self.ui.night.isChecked()
            ):
            error_message = QErrorMessage(self)
            return error_message.showMessage("Seleccione un turno primero (Madrugada a Noche)")
        
        if self.ui.early.isChecked():
            turno = 0
        elif self.ui.morning.isChecked():
            turno = 1
        elif self.ui.evening.isChecked():
            turno = 2
        elif self.ui.night.isChecked():
            turno = 3

        try:
            writing_campo(self.path_file, turno)
        except Exception as e:
            error_message = QErrorMessage(self)
            error_message.showMessage(str(e))

    def data_model(self):
        self.version10 = self.ui.checkBox.isChecked()
        self.version24 = self.ui.checkBox_2.isChecked()
        try:
            if self.version10:
                vissim = com.Dispatch('Vissim.Vissim.10')
            elif self.version24:
                vissim = com.Dispatch('Vissim.Vissim.24')
            else:
                error_message = QErrorMessage(self)
                return error_message.showMessage("Escoger una versión de vissim primero")
        except Exception as inst:
            error_message = QErrorMessage(self)
            return error_message.showMessage("No se pudo conectar al COM")
        
        #Computing veh_classes evaluation:
        try:
            inpx_path = self.path_file
        except Exception as inst:
            error_message = QErrorMessage(self)
            return error_message.showMessage("Selecciona primero la ubicacion del archivo .inpx")
        
        veh_classes_dict, number_vehClasses, nodes_dict = get_veh_classes(inpx_path) #Enviar QErrorMessage Object

        #Open excel
        directory = os.path.dirname(self.path_file)
        modelo = os.path.join(directory, 'Reporte_GEH-R2.xlsm')
        wb = xw.Book(modelo)
        ws = wb.sheets['GEH']

        #Writing data to excel
        nro_row = 8
        total_veh_not_considered = []
        for number_node, node_code in nodes_dict.items():
            try:
                NODE_RES, ORIGIN, DESTINY = get_results(vissim, number_vehClasses, number_node)
            except Exception as inst:
                error_message = QErrorMessage(self)
                return error_message.showMessage(str(inst))
            try:
                count_row, veh_not_considered = writing_excel(
                    NODE_RES = NODE_RES,
                    ORIGIN = ORIGIN,
                    DESTINY = DESTINY,
                    CODE = node_code,
                    veh_classes = veh_classes_dict,
                    ws = ws,
                    nro_row = nro_row,
                )
            except Exception as inst:
                error_message = QErrorMessage(self)
                return error_message.showMessage(str(inst))
            nro_row += count_row + 1 #Para que inicie en al siguiente linea :D
            total_veh_not_considered.extend(veh_not_considered)
        wb.save(modelo)

        total_veh_not_considered = list(set(total_veh_not_considered))
        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.Warning)
        message_box.setWindowTitle("Warning")
        message_box.setText(f"Not considered vehicles: {total_veh_not_considered}")
        return message_box.exec_()
        
    def livianos(self): #Ready
        guide_path = "./images/Parametros_Guia.xlsx"
        wb = load_workbook(guide_path, read_only=True, data_only=True)
        ws = wb['Hoja1']

        self.ui.LookAheadDistMin.setText   (str(ws.cell(5,6).value))
        self.ui.LookAheadDistMax.setText   (str(ws.cell(6,6).value))
        self.ui.LookBackDistMin.setText    (str(ws.cell(8,6).value))
        self.ui.LookBackDistMax.setText    (str(ws.cell(9,6).value))
        self.ui.W74ax.setText              (str(ws.cell(12,6).value))
        self.ui.W74bxAdd.setText           (str(ws.cell(13,6).value))
        self.ui.W74bxMult.setText          (str(ws.cell(14,6).value))
        self.ui.MaxDecelOwn.setText        (str(ws.cell(17,6).value))
        self.ui.MaxDecelTrail.setText      (str(ws.cell(18,6).value))
        self.ui.DecelRedDistOwn.setText    (str(ws.cell(19,6).value))
        self.ui.DecelRedDistTrail.setText  (str(ws.cell(20,6).value))
        self.ui.AccDecelOwn.setText        (str(ws.cell(21,6).value))
        self.ui.AccDecelTrail.setText      (str(ws.cell(22,6).value))
        self.ui.DiffusTm.setText           (str(ws.cell(23,6).value))
        self.ui.SafDistFactLnChg.setText   (str(ws.cell(25,6).value))
        self.ui.CoopDecel.setText          (str(ws.cell(26,6).value))
        self.ui.MinCollTmGain.setText      (str(ws.cell(35,6).value))
        self.ui.MinSpeedForLat.setText     (str(ws.cell(36,6).value))
        self.ui.LatDistStandDef.setText    (str(ws.cell(38,7).value))
        self.ui.LatDistDrivDef.setText     (str(ws.cell(39,7).value))
        wb.close()

    def menores(self): #Ready
        guide_path = "./images/Parametros_Guia.xlsx"
        wb = load_workbook(guide_path, read_only=True, data_only=True)
        ws = wb['Hoja1']

        self.ui.LookAheadDistMin.setText   (str(ws.cell(5,8).value))
        self.ui.LookAheadDistMax.setText   (str(ws.cell(6,8).value))
        self.ui.LookBackDistMin.setText    (str(ws.cell(8,8).value))
        self.ui.LookBackDistMax.setText    (str(ws.cell(9,8).value))
        self.ui.W74ax.setText              (str(ws.cell(12,8).value))
        self.ui.W74bxAdd.setText           (str(ws.cell(13,8).value))
        self.ui.W74bxMult.setText          (str(ws.cell(14,8).value))
        self.ui.MaxDecelOwn.setText        (str(ws.cell(17,8).value))
        self.ui.MaxDecelTrail.setText      (str(ws.cell(18,8).value))
        self.ui.DecelRedDistOwn.setText    (str(ws.cell(19,8).value))
        self.ui.DecelRedDistTrail.setText  (str(ws.cell(20,8).value))
        self.ui.AccDecelOwn.setText        (str(ws.cell(21,8).value))
        self.ui.AccDecelTrail.setText      (str(ws.cell(22,8).value))
        self.ui.DiffusTm.setText           (str(ws.cell(23,8).value))
        self.ui.SafDistFactLnChg.setText   (str(ws.cell(25,8).value))
        self.ui.CoopDecel.setText          (str(ws.cell(26,8).value))
        self.ui.MinCollTmGain.setText      (str(ws.cell(35,8).value))
        self.ui.MinSpeedForLat.setText     (str(ws.cell(36,8).value))
        self.ui.LatDistStandDef.setText    (str(ws.cell(38,9).value))
        self.ui.LatDistDrivDef.setText     (str(ws.cell(39,9).value))
        wb.close()

    def publicos(self): #Ready
        guide_path = "./images/Parametros_Guia.xlsx"
        wb = load_workbook(guide_path, read_only=True, data_only=True)
        ws = wb['Hoja1']

        self.ui.LookAheadDistMin.setText   (str(ws.cell(5,10).value))
        self.ui.LookAheadDistMax.setText   (str(ws.cell(6,10).value))
        self.ui.LookBackDistMin.setText    (str(ws.cell(8,10).value))
        self.ui.LookBackDistMax.setText    (str(ws.cell(9,10).value))
        self.ui.W74ax.setText              (str(ws.cell(12,10).value))
        self.ui.W74bxAdd.setText           (str(ws.cell(13,10).value))
        self.ui.W74bxMult.setText          (str(ws.cell(14,10).value))
        self.ui.MaxDecelOwn.setText        (str(ws.cell(17,10).value))
        self.ui.MaxDecelTrail.setText      (str(ws.cell(18,10).value))
        self.ui.DecelRedDistOwn.setText    (str(ws.cell(19,10).value))
        self.ui.DecelRedDistTrail.setText  (str(ws.cell(20,10).value))
        self.ui.AccDecelOwn.setText        (str(ws.cell(21,10).value))
        self.ui.AccDecelTrail.setText      (str(ws.cell(22,10).value))
        self.ui.DiffusTm.setText           (str(ws.cell(23,10).value))
        self.ui.SafDistFactLnChg.setText   (str(ws.cell(25,10).value))
        self.ui.CoopDecel.setText          (str(ws.cell(26,10).value))
        self.ui.MinCollTmGain.setText      (str(ws.cell(35,10).value))
        self.ui.MinSpeedForLat.setText     (str(ws.cell(36,10).value))
        self.ui.LatDistStandDef.setText    (str(ws.cell(38,11).value))
        self.ui.LatDistDrivDef.setText     (str(ws.cell(39,11).value))
        wb.close()

    def cargas(self): #Ready
        guide_path = "./images/Parametros_Guia.xlsx"
        wb = load_workbook(guide_path, read_only=True, data_only=True)
        ws = wb['Hoja1']

        self.ui.LookAheadDistMin.setText   (str(ws.cell(5,12).value))
        self.ui.LookAheadDistMax.setText   (str(ws.cell(6,12).value))
        self.ui.LookBackDistMin.setText    (str(ws.cell(8,12).value))
        self.ui.LookBackDistMax.setText    (str(ws.cell(9,12).value))
        self.ui.W74ax.setText              (str(ws.cell(12,12).value))
        self.ui.W74bxAdd.setText           (str(ws.cell(13,12).value))
        self.ui.W74bxMult.setText          (str(ws.cell(14,12).value))
        self.ui.MaxDecelOwn.setText        (str(ws.cell(17,12).value))
        self.ui.MaxDecelTrail.setText      (str(ws.cell(18,12).value))
        self.ui.DecelRedDistOwn.setText    (str(ws.cell(19,12).value))
        self.ui.DecelRedDistTrail.setText  (str(ws.cell(20,12).value))
        self.ui.AccDecelOwn.setText        (str(ws.cell(21,12).value))
        self.ui.AccDecelTrail.setText      (str(ws.cell(22,12).value))
        self.ui.DiffusTm.setText           (str(ws.cell(23,12).value))
        self.ui.SafDistFactLnChg.setText   (str(ws.cell(25,12).value))
        self.ui.CoopDecel.setText          (str(ws.cell(26,12).value))
        self.ui.MinCollTmGain.setText      (str(ws.cell(35,12).value))
        self.ui.MinSpeedForLat.setText     (str(ws.cell(36,12).value))
        self.ui.LatDistStandDef.setText    (str(ws.cell(38,13).value))
        self.ui.LatDistDrivDef.setText     (str(ws.cell(39,13).value))
        wb.close()

    def get(self): #Ready
        self.version10 = self.ui.checkBox.isChecked()
        self.version24 = self.ui.checkBox_2.isChecked()
        #INICIO DE COM
    
        if self.version10:
            try:
                vissim = com.Dispatch('Vissim.Vissim.10')
            except Exception as e:
                error_message = QErrorMessage(self)
                return error_message.showMessage("No se pudo conectar al COM")

        elif self.version24:
            try:
                vissim = com.Dispatch('Vissim.Vissim.24')
            except Exception as e:
                error_message = QErrorMessage(self)
                return error_message.showMessage("No se pudo conectar al COM")
        else:
            error_message = QErrorMessage(self)
            return error_message.showMessage("Escoger una versión de vissim primero")

        key = int(self.ui.vehicle_type.text())
        
        self.ui.LookAheadDistMin.setText       (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('LookAheadDistMin')))
        self.ui.LookAheadDistMax.setText       (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('LookAheadDistMax')))
        self.ui.LookBackDistMin.setText        (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('LookBackDistMin')))
        self.ui.LookBackDistMax.setText        (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('LookBackDistMax')))
        self.ui.W74ax.setText                  (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('W74ax')))
        self.ui.W74bxAdd.setText               (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('W74bxAdd')))
        self.ui.W74bxMult.setText              (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('W74bxMult')))
        self.ui.MaxDecelOwn.setText            (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('MaxDecelOwn')))
        self.ui.MaxDecelTrail.setText          (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('MaxDecelTrail')))
        self.ui.DecelRedDistOwn.setText        (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('DecelRedDistOwn')))
        self.ui.DecelRedDistTrail.setText      (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('DecelRedDistTrail')))
        self.ui.AccDecelOwn.setText            (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('AccDecelOwn')))
        self.ui.AccDecelTrail.setText          (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('AccDecelTrail')))
        self.ui.DiffusTm.setText               (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('DiffusTm')))
        self.ui.SafDistFactLnChg.setText       (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('SafDistFactLnChg')))
        self.ui.CoopDecel.setText              (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('CoopDecel')))
        self.ui.MinCollTmGain.setText          (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('MinCollTmGain')))
        self.ui.MinSpeedForLat.setText         (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('MinSpeedForLat')))
        self.ui.LatDistStandDef.setText        (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('LatDistStandDef')))
        self.ui.LatDistDrivDef.setText         (str(vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('LatDistDrivDef')))

        DesLatPos_ui = vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('DesLatPos')
        if DesLatPos_ui == 'MIDDLE':
            self.ui.DesLatPos.setCurrentIndex(0)
        elif DesLatPos_ui == 'ANY':
            self.ui.DesLatPos.setCurrentIndex(1)
        elif DesLatPos_ui == 'RIGHT':
            self.ui.DesLatPos.setCurrentIndex(2)
        elif DesLatPos_ui == 'LEFT':
            self.ui.DesLatPos.setCurrentIndex(3)

        if vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('ObsrvAdjLn') == 0:
            self.ui.ObsrvAdjLn.setChecked(False)
        else: self.ui.ObsrvAdjLn.setChecked(True)
        if vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('DiamQueu') == 0:
            self.ui.DiamQueu.setChecked(False)
        else: self.ui.DiamQueu.setChecked(True)
        if vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('ConsNextTurn') == 0:
            self.ui.ConsNextTurn.setChecked(False)
        else: self.ui.ConsNextTurn.setChecked(True)
        if vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('OvtLDef') == 0:
            self.ui.OvtLDef.setChecked(False)
        else: self.ui.OvtLDef.setChecked(True)
        if vissim.Net.DrivingBehaviors.ItemByKey(key).AttValue('OvtRDef') == 0:
            self.ui.OvtRDef.setChecked(False)
        else: self.ui.OvtRDef.setChecked(True)        

    def ejecutar_programa(self): #Ready
        global enviado_contador
        ####INTRODUCCION DE DATA####
        vehicle_type        =self.ui.vehicle_type.text()
        LookAheadDistMin    =self.ui.LookAheadDistMin.text()
        LookAheadDistMax    =self.ui.LookAheadDistMax.text()
        LookBackDistMin     =self.ui.LookBackDistMin.text()
        LookBackDistMax     =self.ui.LookBackDistMax.text()

        ####WIEDEMANN 74####
        W74ax               =self.ui.W74ax.text()
        W74bxAdd            =self.ui.W74bxAdd.text()
        W74bxMult           =self.ui.W74bxMult.text()

        ####CAMBIO DE CARRIL####
        MaxDecelOwn         =self.ui.MaxDecelOwn.text()
        MaxDecelTrail       =self.ui.MaxDecelTrail.text()
        DecelRedDistOwn     =self.ui.DecelRedDistOwn.text()
        DecelRedDistTrail   =self.ui.DecelRedDistTrail.text()
        AccDecelOwn         =self.ui.AccDecelOwn.text()
        AccDecelTrail       =self.ui.AccDecelTrail.text()
        DiffusTm            =self.ui.DiffusTm.text()
        SafDistFactLnChg    =self.ui.SafDistFactLnChg.text()
        CoopDecel           =self.ui.CoopDecel.text()

        ####LATERAL####
        MinCollTmGain       =self.ui.MinCollTmGain.text()
        MinSpeedForLat      =self.ui.MinSpeedForLat.text()
        LatDistStandDef     =self.ui.LatDistStandDef.text()
        LatDistDrivDef      =self.ui.LatDistDrivDef.text()

        ####ACCESORIES####
        DesLatPos           =self.ui.DesLatPos.currentText()

        if self.ui.ObsrvAdjLn.isChecked(): ObsrvAdjLn="true"
        else: ObsrvAdjLn="false"

        if self.ui.DiamQueu.isChecked(): DiamQueu="true"
        else: DiamQueu="false"
  
        if self.ui.ConsNextTurn.isChecked(): ConsNextTurn="true"
        else: ConsNextTurn="false"

        if self.ui.OvtLDef.isChecked(): OvtLDef="true"
        else: OvtLDef="false"

        if self.ui.OvtRDef.isChecked(): OvtRDef="true"
        else: OvtRDef="false"

        ####NEWS####
        try:
            if self.version24:
                if self.ui.checkBox_3.isChecked():
                    Zipper              = "true"
                    ZipperMinSpeed      = self.ui.ZipperMinSpeed.text()
        except Exception as e:
            error_message = QErrorMessage(self)
            return error_message.showMessage("Selecciona la versión de vissim primero.")

        #INICIO DE COM
        try:
            if self.version10:
                vissim = com.Dispatch('Vissim.Vissim.10')
            elif self.version24:
                vissim = com.Dispatch('Vissim.Vissim.24')
        except Exception as e:
            error_message = QErrorMessage(self)
            return error_message.showMessage("No hay una estancia de Vissim activa o no has seleccionado una versión.")

        if vissim.Simulation.AttValue('IsRunning'):
                vissim.Simulation.RunSingleStep() #Pause

        ##################################
        # ENVIO DE INFORMACIÓN AL VISSIM #
        ##################################

        try:
            key = int(vehicle_type)
        except Exception as inst:
            error_message = QErrorMessage(self)
            return error_message.showMessage("Escoge el 'No' del comportamiento vehicular")

        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("LookAheadDistMin",  LookAheadDistMin)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("LookAheadDistMax",  LookAheadDistMax)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("LookBackDistMin",   LookBackDistMin)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("LookBackDistMax",   LookBackDistMax)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("W74ax",             W74ax)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("W74bxAdd",          W74bxAdd)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("W74bxMult",         W74bxMult)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("MaxDecelOwn",       MaxDecelOwn)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("MaxDecelTrail",     MaxDecelTrail)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("DecelRedDistOwn",   DecelRedDistOwn)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("DecelRedDistTrail", DecelRedDistTrail)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("AccDecelOwn",       AccDecelOwn)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("AccDecelTrail",     AccDecelTrail)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("DiffusTm",          DiffusTm)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("SafDistFactLnChg",  SafDistFactLnChg)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("CoopDecel",         CoopDecel)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("MinCollTmGain",     MinCollTmGain)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("MinSpeedForLat",    MinSpeedForLat)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("LatDistStandDef",   LatDistStandDef)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("LatDistDrivDef",    LatDistDrivDef)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("DesLatPos",         DesLatPos)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("ObsrvAdjLn",        ObsrvAdjLn)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("DiamQueu",          DiamQueu)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("ConsNextTurn",      ConsNextTurn)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("OvtLDef",           OvtLDef)
        vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("OvtRDef",           OvtRDef)
        #ZIPPER:
        if self.version24:
            if self.ui.checkBox_3.isChecked():
                vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("Zipper",Zipper)
                vissim.Net.DrivingBehaviors.ItemByKey(key).SetAttValue("ZipperMinSpeed",ZipperMinSpeed)

        #CHECK
        self.ui.enviado.setText(f"ENVIADO {enviado_contador}")
        enviado_contador += 1

    def fijars(self): #Ready
        global param_groups

        if self.ui.vehicle_type.text() == '':
            error_message = QErrorMessage(self)
            return error_message.showMessage("Ingresa el comportamiento vehicular primero.")

        default_dictionary = {
            1: self.ui.LookAheadDistMin.text(),
            2: self.ui.LookAheadDistMax.text(),
            3: self.ui.LookBackDistMin.text(),
            4: self.ui.LookBackDistMax.text(),
            5: self.ui.W74ax.text(),
            6: self.ui.W74bxAdd.text(),
            7: self.ui.W74bxMult.text(),
            8: self.ui.MaxDecelOwn.text(),
            9: self.ui.MaxDecelTrail.text(),
            10: self.ui.DecelRedDistOwn.text(),
            11: self.ui.DecelRedDistTrail.text(),
            12: self.ui.AccDecelOwn.text(),
            13: self.ui.AccDecelTrail.text(),
            14: self.ui.DiffusTm.text(),
            15: self.ui.SafDistFactLnChg.text(),
            16: self.ui.CoopDecel.text(),
            17: self.ui.DesLatPos.currentText(),
            18: str(self.ui.ObsrvAdjLn.isChecked()),
            19: str(self.ui.DiamQueu.isChecked()),
            20: str(self.ui.ConsNextTurn.isChecked()),
            21: self.ui.MinCollTmGain.text(),
            22: self.ui.MinSpeedForLat.text(),
            23: str(self.ui.OvtLDef.isChecked()),
            24: str(self.ui.OvtRDef.isChecked()),
            25: self.ui.LatDistStandDef.text(),
            26: self.ui.LatDistDrivDef.text(),
        }

        param_groups[self.ui.vehicle_type.text()] = default_dictionary

        self.ui.enviado.setText(f"Send {self.ui.vehicle_type.text()}")

    def export_params_2_excel(self): #Ready
        global param_groups
        original_path = "./images/Parametros_Model.xlsx"
        try:
            directorio,_ = os.path.split(self.path_file)
        except AttributeError:
            error_message = QErrorMessage(self)
            return error_message.showMessage("Debes seleccionar la ubicación del archivo Vissim para que se guarde allí el excel.")
        
        modelo = os.path.join(directorio,'Parametros_Calibracion.xlsx')

        shutil.copy2(original_path,modelo)
        workbook = load_workbook(modelo)
        worksheet = workbook['Hoja1']

        ######################
        # INGRESO DE VALORES #
        ######################
        files = [5,6,8,9,12,13,14,17,18,19,20,21,22,23,25,26,31,32,33,34,35,36,38,39,38,39]
        
        for index1,(_,group) in enumerate(param_groups.items()):
            worksheet.cell(4,6+index1*2).value = int(index1+1)
            for i,(_,value) in enumerate(group.items()):
                if i<24:
                    try:
                        worksheet.cell(files[i],6+index1*2).value=float(value)
                    except ValueError:
                        worksheet.cell(files[i],6+index1*2).value=str(value)
                elif i==24:
                    try:
                        worksheet.cell(files[i],7+index1*2).value=float(value)
                    except ValueError:
                        worksheet.cell(files[i],7+index1*2).value=str(value)
                else:
                    try:
                        worksheet.cell(files[i],7+index1*2).value=float(value)
                    except ValueError:
                        worksheet.cell(files[i],7+index1*2).value=str(value)

        workbook.save(modelo)
        workbook.close()
        self.ui.enviado.setText("PÁRAMETROS OK!")

    def run_vissim(self):
        numruns = self.ui.spinBox.value()
        simres = self.ui.spinBox_2.value()
        numcores = self.ui.spinBox_3.value()

        self.version10 = self.ui.checkBox.isChecked()
        self.version24 = self.ui.checkBox_2.isChecked()
        try:
            if self.version10:
                vissim = com.Dispatch('Vissim.Vissim.10')
            elif self.version24:
                vissim = com.Dispatch('Vissim.Vissim.24')
            else:
                error_message = QErrorMessage(self)
                return error_message.showMessage("Escoger una versión de vissim primero")
        except Exception as inst:
            error_message = QErrorMessage(self)
            return error_message.showMessage("No se pudo conectar al COM")
        
        vissim.Simulation.SetAttValue("NumRuns", numruns)
        vissim.Simulation.SetAttValue("SimRes", simres)
        vissim.Simulation.SetAttValue("NumCores", numcores)
        vissim.Simulation.SetAttValue("SimPeriod", 5400)
        vissim.Evaluation.SetAttValue("NodeResCollectData", True)
        vissim.Evaluation.SetAttValue("NodeResToTime", 5400)
        vissim.Evaluation.SetAttValue("NodeResFromTime", 1800)
        vissim.Evaluation.SetAttValue("NodeResInterval", 3600)
        vissim.Evaluation.SetAttValue("VehNetPerfCollectData", False)
        vissim.Evaluation.SetAttValue("PedNetPerfCollectData", False)
        vissim.Graphics.SetAttValue("QuickMode", True)
        vissim.Simulation.RunContinuous()

def main():
    app = QApplication([])
    window = MiVentana()
    window.show()
    app.exec_()

if __name__ == "__main__": 
    main()