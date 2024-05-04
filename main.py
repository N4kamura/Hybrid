from PyQt5 import uic
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QButtonGroup, QErrorMessage, QMessageBox
from PyQt5.QtGui import QPixmap
import win32com.client as com
from pywintypes import com_error
import os
from openpyxl import load_workbook
import shutil
import warnings
from writing import writing_campo
from openpyxl import load_workbook
from hybrid_ui import Ui_MainWindow
import xlwings as xw
#Customize modules
from model import *
from guides.tools import *
from vissim.get_info import *
from vissim.send_info import *
from excels.get_info import *

warnings.filterwarnings("ignore", category=DeprecationWarning)
nombres_vehiculos = [] #TODO: Implementar a futuro
dict_behaviors = {}

class MiVentana(QMainWindow):
    def __init__(self): #Ready
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        ####IMAGES####
        imagen = QPixmap('./images/car_follow.png')
        self.ui.label.setPixmap(imagen)
        imagen_2 = QPixmap('./images/lane_change.png')
        self.ui.label_2.setPixmap(imagen_2)
        image_3 = QPixmap('./images/lateral_behave.png')
        self.ui.label_3.setPixmap(image_3)
        image_4 = QPixmap('./images/logo.png')
        self.ui.label_5.setPixmap(image_4)

        #BOTONES

        self.ui.start.clicked.connect(self.ejecutar_programa)
        self.ui.carpet.clicked.connect(self.carpeta)
        self.ui.report.clicked.connect(self.data_model)
        self.ui.activar.clicked.connect(self.data_campo)
        self.ui.liviano.clicked.connect(self.livianos)
        self.ui.menor.clicked.connect(self.menores)
        self.ui.publico.clicked.connect(self.publicos)
        self.ui.carga.clicked.connect(self.cargas)
        self.ui.fijar.clicked.connect(self.fijars)
        self.ui.exportar.clicked.connect(self.export_params_2_excel)
        self.ui.get_pushButton.clicked.connect(self.get)
        self.ui.run.clicked.connect(self.run_vissim)
        
        #BOTONES PARA LOS TURNOS
        button_group = QButtonGroup(self)
        button_group.addButton(self.ui.early)
        button_group.addButton(self.ui.morning)
        button_group.addButton(self.ui.evening)
        button_group.addButton(self.ui.night)
        button_group.setExclusive(True)

        #BOTONES PARA LAS VERSIONES
        button_group_2 = QButtonGroup(self)
        button_group_2.addButton(self.ui.checkBox)
        button_group_2.addButton(self.ui.checkBox_2)

    def carpeta(self): #OK
        self.path_file, self.inpx_name = QFileDialog.getOpenFileName(self, "Seleccionar Archivo .inpx","c:\\","Archivos .inpx (*.inpx)")

    def data_campo(self): #OK
        #Checking buttons:
        while not (
            self.ui.early.isChecked() or
            self.ui.morning.isChecked() or
            self.ui.evening.isChecked() or
            self.ui.night.isChecked()
            ):
            error_message = QErrorMessage(self)
            return error_message.showMessage("Seleccione un turno primero (Madrugada a Noche)")
        
        if self.ui.early.isChecked():
            turno = 0
        elif self.ui.morning.isChecked():
            turno = 1
        elif self.ui.evening.isChecked():
            turno = 2
        elif self.ui.night.isChecked():
            turno = 3

        try:
            writing_campo(self.path_file, turno)
        except Exception as e:
            error_message = QErrorMessage(self)
            error_message.showMessage(str(e))

    def data_model(self): #REWORKING
        self.version10 = self.ui.checkBox.isChecked()
        self.version24 = self.ui.checkBox_2.isChecked()
        try:
            if self.version10:
                vissim = com.Dispatch('Vissim.Vissim.10')
            elif self.version24:
                vissim = com.Dispatch('Vissim.Vissim.24')
            else:
                error_message = QErrorMessage(self)
                return error_message.showMessage("Escoger una versión de vissim primero")
        except Exception as inst:
            error_message = QErrorMessage(self)
            return error_message.showMessage("No se pudo conectar al COM")
        
        #Computing veh_classes evaluation:
        try:
            inpx_path = self.path_file
        except Exception as inst:
            error_message = QErrorMessage(self)
            return error_message.showMessage("Selecciona primero la ubicacion del archivo .inpx")
        
        veh_classes_dict, number_vehClasses, nodes_dict = get_veh_classes(inpx_path) #Enviar QErrorMessage Object

        #Open excel
        directory = os.path.dirname(self.path_file)
        modelo = os.path.join(directory, 'Reporte_GEH-R2.xlsm')
        wb = xw.Book(modelo)
        ws = wb.sheets['GEH']

        #Writing data to excel
        nro_row = 8
        total_veh_not_considered = []
        for number_node, node_code in nodes_dict.items():
            try:
                NODE_RES, ORIGIN, DESTINY = get_results(vissim, number_vehClasses, number_node)
            except Exception as inst:
                error_message = QErrorMessage(self)
                return error_message.showMessage(str(inst))
            try:
                count_row, veh_not_considered = writing_excel(
                    NODE_RES = NODE_RES,
                    ORIGIN = ORIGIN,
                    DESTINY = DESTINY,
                    CODE = node_code,
                    veh_classes = veh_classes_dict,
                    ws = ws,
                    nro_row = nro_row,
                )
            except Exception as inst:
                error_message = QErrorMessage(self)
                return error_message.showMessage(str(inst))
            nro_row += count_row #Para que inicie en al siguiente linea :D
            total_veh_not_considered.extend(veh_not_considered)

        wb.save(modelo)
        total_veh_not_considered = list(set(total_veh_not_considered))
        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.Warning)
        message_box.setWindowTitle("Warning")
        message_box.setText(f"Not considered vehicles: {total_veh_not_considered}")
        return message_box.exec_()
        
    def livianos(self): #OK
        guide_path = "./images/Parametros_Guia.xlsx"
        send_light(self.ui, guide_path)

    def menores(self): #OK
        guide_path = "./images/Parametros_Guia.xlsx"
        send_minors(self.ui, guide_path)

    def publicos(self): #OK
        guide_path = "./images/Parametros_Guia.xlsx"
        send_public(self.ui, guide_path)

    def cargas(self): #OK
        guide_path = "./images/Parametros_Guia.xlsx"
        send_hgv(self.ui, guide_path)

    def get(self): #OK
        #INICIO DE COM

        if self.ui.checkBox.isChecked():
            version = 10
        elif self.ui.checkBox_2.isChecked():
            version = 24
        else:
            error_message = QErrorMessage(self)
            return error_message.showMessage("Escoger una versión de vissim primero")
    
        try:
            vissim = com.Dispatch(f'Vissim.Vissim.{version}')
        except Exception as e:
            error_message = QErrorMessage(self)
            return error_message.showMessage("No se pudo conectar al COM")
        
        try:
            error_message = QErrorMessage(self)
            show_info(vissim, self.ui, version, error_message)
        except ValueError:
            return error_message.showMessage("Selecciona un número de comportamiento vehicular")

    def ejecutar_programa(self): #OK
        if self.ui.checkBox_2.isChecked(): #v24
            DATA = from_ui_24(self.ui)
            version = 24
        elif self.ui.checkBox.isChecked(): #v10
            DATA = from_ui_10(self.ui)
            version = 10
        else:
            error_message = QErrorMessage(self)
            return error_message.showMessage("Escoge la versión de vissim.")

        #INICIO DE COM
        try:
            vissim = com.Dispatch(f'Vissim.Vissim.{version}')
        except Exception as e:
            error_message = QErrorMessage(self)
            return error_message.showMessage("No hay una estancia de Vissim COM activa")

        #PAUSE
        if vissim.Simulation.AttValue('IsRunning'):
            vissim.Simulation.RunSingleStep() #Pause

        ##################################
        # ENVIO DE INFORMACIÓN AL VISSIM #
        ##################################

        try:
            key = int(DATA.vehicleType)
        except Exception as inst:
            error_message = QErrorMessage(self)
            return error_message.showMessage("Escoge el 'No' del comportamiento vehicular")
        
        behavior = vissim.Net.DrivingBehaviors.ItemByKey(key)

        if version == 24:
            to_24(behavior, DATA)
        elif version == 10:
            to_10(behavior, DATA)

        return self.ui.enviado(f"Enviado Nro. {key}")

    def fijars(self): #OK
        if self.ui.vehicle_type.text() == '':
            error_message = QErrorMessage(self)
            return error_message.showMessage("Usa el formato 1-4,5,7-11 por ejemplo.")
        
        try:
            no_behaviors = parse_numbers(self.ui.vehicle_type.text())
        except (ValueError, TypeError):
            error_message = QErrorMessage(self)
            return error_message.showMessage("Usa el formato 1-4,5,7-11 por ejemplo.")
        
        if self.ui.checkBox.isChecked():
            version = 10
        elif self.ui.checkBox_2.isChecked():
            version = 24
        else:
            error_message = QErrorMessage(self)
            return error_message.showMessage("Escoger una versión de vissim primero")
        
        try:
            vissim = com.Dispatch(f'Vissim.Vissim.{version}')
        except Exception as e:
            error_message = QErrorMessage(self)
            return error_message.showMessage("No se pudo conectar al COM")

        for nro_behavior in no_behaviors:
            if nro_behavior <= 0: continue

            try:
                v = vissim.Net.DrivingBehaviors.ItemByKey(nro_behavior)
            except com_error as e:
                error_message = QErrorMessage(self)
                return error_message.showMessage(f"El comportamiento {nro_behavior} no existe")
            
            if version == 24:
                data = extractor_24(v)
            elif version == 10:
                data = extractor_10(v)
            
            dict_behaviors[nro_behavior] = get_variables24(data)

        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.Information)
        message_box.setWindowTitle("INFO")
        message_box.setText(f"¡Comportamientos guardados: {no_behaviors}!")
        return message_box.exec_()

    def export_params_2_excel(self): #OK
        original_path = "./images/Parametros_Model.xlsx"
        try:
            directorio,_ = os.path.split(self.path_file)
        except AttributeError:
            error_message = QErrorMessage(self)
            return error_message.showMessage("Debes seleccionar la ubicación del archivo Vissim para que se guarde allí el excel.")
        
        modelo = os.path.join(directorio,'Parametros_Calibracion.xlsx')

        shutil.copy2(original_path,modelo)
        workbook = load_workbook(modelo)
        worksheet = workbook['Hoja1']

        ######################
        # INGRESO DE VALORES #
        ######################
        files = [5,6,8,9,12,13,14,17,18,19,20,21,22,23,25,26,31,32,33,34,35,36,38,39,38,39]
        
        for index1,(_,group) in enumerate(dict_behaviors.items()):
            worksheet.cell(4,6+index1*2).value = int(index1+1)
            for i,(_,value) in enumerate(group.items()):
                if i<24:
                    try:
                        worksheet.cell(files[i],6+index1*2).value=float(value)
                    except ValueError:
                        worksheet.cell(files[i],6+index1*2).value=str(value)
                elif i==24:
                    try:
                        worksheet.cell(files[i],7+index1*2).value=float(value)
                    except ValueError:
                        worksheet.cell(files[i],7+index1*2).value=str(value)
                else:
                    try:
                        worksheet.cell(files[i],7+index1*2).value=float(value)
                    except ValueError:
                        worksheet.cell(files[i],7+index1*2).value=str(value)

        workbook.save(modelo)
        workbook.close()
        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.Information)
        message_box.setWindowTitle("INFO")
        message_box.setText(f"¡Se creo un excel con los parámetros de comportamiento exitosamente!")
        return message_box.exec_()

    def run_vissim(self): #OK
        numruns = self.ui.spinBox.value()
        simres = self.ui.spinBox_2.value()
        numcores = self.ui.spinBox_3.value()

        self.version10 = self.ui.checkBox.isChecked()
        self.version24 = self.ui.checkBox_2.isChecked()
        try:
            if self.version10:
                vissim = com.Dispatch('Vissim.Vissim.10')
            elif self.version24:
                vissim = com.Dispatch('Vissim.Vissim.24')
            else:
                error_message = QErrorMessage(self)
                return error_message.showMessage("Escoger una versión de vissim primero")
        except Exception as inst:
            error_message = QErrorMessage(self)
            return error_message.showMessage("No se pudo conectar al COM")
        

        try:
            script_path = os.path.dirname(os.path.abspath(__file__))
            layout_path = os.path.join(script_path,"images","layout.layx")
            vissim.loadLayout(layout_path)
        except com_error as inst:
            error_message = QErrorMessage(self)
            return error_message.showMessage("No se pudo cargar el archivo de la red, revisar si lo tienes en la carpeta images/layout.layx")
        
        vissim.Simulation.SetAttValue("NumRuns", numruns)
        vissim.Simulation.SetAttValue("SimRes", simres)
        vissim.Simulation.SetAttValue("NumCores", numcores)
        vissim.Simulation.SetAttValue("SimPeriod", 5400)
        vissim.Evaluation.SetAttValue("NodeResCollectData", True)
        vissim.Evaluation.SetAttValue("NodeResToTime", 5400)
        vissim.Evaluation.SetAttValue("NodeResFromTime", 1800)
        vissim.Evaluation.SetAttValue("NodeResInterval", 3600)
        vissim.Evaluation.SetAttValue("VehNetPerfCollectData", False)
        vissim.Evaluation.SetAttValue("PedNetPerfCollectData", False)
        vissim.Graphics.SetAttValue("QuickMode", True)
        vissim.Simulation.RunContinuous()

def main():
    app = QApplication([])
    window = MiVentana()
    window.show()
    app.exec_()

if __name__ == "__main__": 
    main()