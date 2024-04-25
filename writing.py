import xml.etree.ElementTree as ET
import xlwings as xw
import os
from collections import Counter
import shutil
from openpyxl import load_workbook
from tools import read_one_excel, peakhourfinder, obtener_numero_al_final
from unidecode import unidecode

def writing_campo(vissim_path, turno) -> None:
    tree = ET.parse(vissim_path)
    network_tag = tree.getroot()

    uda_key = network_tag.find(".//userDefinedAttribute[@nameLong='Código']")
    key = uda_key.get('no')
    nodes_info = []

    for node_tag in network_tag.findall("./nodes/node"):
        number_node = node_tag.get("no")
        uda_element = node_tag.find(f"./uda[@key='{key}']")
        codigo_intersection = uda_element.get("value")
        nodes_info.append((number_node,codigo_intersection))

    nombre_subarea = vissim_path
    for _ in range(4):
        nombre_subarea = os.path.dirname(nombre_subarea)
    _,nombre_subarea = os.path.split(nombre_subarea)

    tipicidad = vissim_path
    for _ in range(2):
        tipicidad = os.path.dirname(tipicidad)
    _,tipicidad = os.path.split(tipicidad)

    directorio_proyecto = vissim_path
    for _ in range(6):
        directorio_proyecto = os.path.dirname(directorio_proyecto)

    directorio_flujogramas = os.path.join(directorio_proyecto, "7. Informacion de Campo", nombre_subarea, "Vehicular", tipicidad)

    files = os.listdir(directorio_flujogramas)

    flujogramas_ordered = []

    for node_info in nodes_info:
        for file in files:
            if file[:5] == node_info[1]: #TODO: <------ Se puede mejorar solo captando patrones.
                flujogramas_ordered.append(file)
                break

    excel_paths = []
    for flujograma in flujogramas_ordered:
        flujograma = os.path.join(directorio_flujogramas,flujograma)
        excel_paths.append(flujograma)
    
    #######################################
    # GETTING NAMES AND NUMBER OF CLASSES #
    #######################################

    wb = load_workbook(excel_paths[0], read_only=True, data_only=True)
    ws = wb['Inicio']
    try:
        num_veh_classes = [row[0].value for row in ws[slice("AD4","AD23")]].index("n")
    except Exception as _:
        num_veh_classes = 20
    types = [unidecode(row[0].value).upper() for row in ws[slice("AD4","AD23")]][:num_veh_classes]
    wb.close()

    #####################
    # FINDING PEAK HOUR #
    #####################

    resultados_peakhour = []

    for excel_path in excel_paths:
        resultados_peakhour.append(peakhourfinder(turno, excel_path))

    horas_puntas = []

    horas_puntas = [hora[0] for hora in resultados_peakhour]
    counter_horas = Counter(horas_puntas)
    modas_horas = [hora for hora, frecuencia in counter_horas.items() if frecuencia == max(counter_horas.values())]

    mayor_volumen = 0

    for moda_hora in modas_horas:
        volumenes_moda = [hora[1] for hora in resultados_peakhour if hora[0] == moda_hora]
        max_volumen_moda = max(volumenes_moda)

        if max_volumen_moda > mayor_volumen:
            mayor_volumen = max_volumen_moda
            peakhour = moda_hora

    ##################
    # READING EXCELS #
    ##################

    interval = slice(peakhour, peakhour+4)
    data_excel = []
    for excel_path in excel_paths:
        data_excel.append(read_one_excel(excel_path, num_veh_classes, interval))

    #Copia del modelo.
    modelo_original = "./images/Modelo.xlsm"
    directorio,_ = os.path.split(vissim_path)
    modelo = os.path.join(directorio,'Reporte_GEH-R2.xlsm')
    shutil.copy2(modelo_original,modelo)

    vehicles_names = 'H8'
    intersection = 'B8'
    od = 'C8'
    vehicle_type = 'D8'
    volumes = 'E8'

    #############################
    # Writing field data in GEH #
    #############################

    wb = xw.Book(modelo)
    ws = wb.sheets['GEH']
    all_values = []

    for i in range(len(types)):
        ws.range(vehicles_names).value = types[i]
        vehicles_names = vehicles_names[0] + str(9+i)

    for excel in data_excel:
        for access in excel:
            for i, turns in enumerate(access[1]):
                for j, volume in enumerate(access[2]):
                    all_values.append([access[0], turns, types[j], volume[i]])

    for index, values in enumerate(all_values):
        ws.range(intersection).value = values[0]
        ws.range(od).value = values[1]
        ws.range(vehicle_type).value = values[2]
        ws.range(volumes).value = values[3]

        intersection = intersection[0] + str(9+index)
        od = od[0] + str(9+index)
        vehicle_type = vehicle_type[0] + str(9+index)
        volumes = volumes[0] + str(9+index) 

    wb.save(modelo)
    wb.close()
    xw.App().quit()