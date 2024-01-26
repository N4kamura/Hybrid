import re
from openpyxl import load_workbook
import numpy as np
import os

def obtener_numero_al_final(archivo):
    match = re.search(r'(\d+)\.att$', archivo)
    return int(match.group(1)) if match else 0

def peakhourfinder(args): 
    turno,excel_path = args
    slices = [
            ["HR16:HR39",0+1,"MADRUGADA"], #El +1 es para que obtenga el límite superior, p.e.: [8:00 - 8:15] -> 8:15
            ["HR40:HR63",24+1,"MAÑANA"],
            ["HR64:HR83",48+1,"TARDE"],
            ["HR84:HR111",68+1,"NOCHE"],
            ]
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["N"]

    list = [[cell.value for cell in row] for row in ws[slices[turno][0]]]
    maximum = list[0]
    maximum_cell = 0
    for i in range(1,len(list)):
        if list[i]>maximum:
            maximum         = list[i]
            maximum_cell    = i

    if maximum[0] !=0: #Acá obtengo el hora inicio. p.e.: 9*4 a 10*4 (9-10) o 9.25 a 10.25 (9:15 - 10:15)
        plus = slices[turno][1]
        peakhour = maximum_cell + plus
        hora_final, hora_inicio = peakhour,peakhour-4
        hours = int(peakhour//4)
        minutes = int(((peakhour/4)%1)*100*0.15/0.25)
        #print(f"Excel: {excel_path}")
        print(f"PEAK HOUR {slices[turno][2]}: {hours-1:02d}:{minutes:02d} - {hours:02d}:{minutes:02d} | Volumen = {maximum[0]}")
    else:
        print(f"PEAK HOUR {slices[turno][2]}: NO HAY DATOS O EL FLUJO ES NULO")
        pass
    
    return hora_inicio,hora_final, maximum

def read_one_excel(excel_path,num_veh_classes,interval):
    wb = load_workbook(excel_path,read_only=True,data_only=True)

    ws_inicio = wb['Inicio']

    num_giros = [
        [row[0].value for row in ws_inicio[s]].index(None)
        for s in [
            slice("G12", "G21"),
            slice("M12", "M21"),
            slice("G24", "G33"),
            slice("M24", "M33"),
        ]    
    ]

    name_giros = ["N", "S", "E", "O"]

    list_destination = []
    list_origin      = []

    list_slice_destination = (
        slice("F12", "F21"),
        slice("L12", "L21"),
        slice("F24", "F33"),
        slice("L24", "L33"),
    )

    list_slice_origin = (
        slice("E12", "E21"),
        slice("K12", "K21"),
        slice("E24", "E33"),
        slice("K24", "K33"),
    )
    final_result=[]
    for i_giro in range(4):
        list_destination = []
        list_origin      = []
        slice_origin = list_slice_origin[i_giro]
        slice_destination = list_slice_destination[i_giro]
        num_giro_i = num_giros[i_giro]

        list_destination.append([row[0].value for row in ws_inicio[slice_destination]][:num_giro_i])

        list_origin.append([row[0].value for row in ws_inicio[slice_origin]][:num_giro_i])

        list_od=[[str(a)+str(b) for a,b in zip(fila1,fila2)] for fila1, fila2 in zip(list_origin,list_destination)]
        list_od = list_od[0]

        ###### GET FLOWS ######
        ws = wb[name_giros[i_giro]]
        list_A = [[cell.value for cell in row] for row in ws["K16":"HB111"]]

        #Conversión a numpy.
        A = np.array(list_A, dtype="float")

        #Asignación de 0 para datos que no contienen data.
        A[np.isnan(A)] = 0

        list_sumas=[]
        matriz = np.array([A[interval, (10 * veh_type) : (10 * veh_type + num_giro_i)] for veh_type in range(num_veh_classes)])
        for mat in matriz:
            list_sumas.append([sum(mat[:,i]) for i in range(num_giro_i)]) #Saldría la suma de todos los giros en una lista y todo esto por cada tipo vehicular.
        
        _, name_excel = os.path.split(excel_path)
        final_result.append([name_excel[:-5],list_od,list_sumas])
    
    wb.close()

    return final_result